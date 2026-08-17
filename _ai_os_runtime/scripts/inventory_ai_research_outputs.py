#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import html
import csv
import json
import mimetypes
import os
import re
import subprocess
import zipfile
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional runtime dependency
    PdfReader = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional runtime dependency
    load_workbook = None


SOURCE_SYSTEM_NAME = "AI generated research outputs"
SOURCE_SYSTEM_LOCATION = "local research output folders and standalone dashboard/report files"

DEFAULT_SOURCE_ROOTS = [
    ("cowork_research", Path("/Users/devarshthakkar/Downloads/cowork reseaarch")),
    ("cowork_outputs", Path("/Users/devarshthakkar/Downloads/cowork outputs")),
    ("claude_cowork_outputs", Path("/Users/devarshthakkar/Downloads/claude cowork outputs")),
    ("claude_outputs", Path("/Users/devarshthakkar/Downloads/claude outputs")),
    ("claude_outputs", Path("/Users/devarshthakkar/Downloads/Claude outputs")),
    ("claude_documents", Path("/Users/devarshthakkar/Documents/Claude")),
    ("codex_outputs", Path("/Users/devarshthakkar/Downloads/codex outputs")),
    ("desktop_codex_outputs", Path("/Users/devarshthakkar/Desktop/codex outputs")),
    ("ultimate_foils", Path("/Users/devarshthakkar/Downloads/ultimate foils data")),
]

def configured_source_roots() -> list[tuple[str, Path]]:
    """Return bounded local and shared inboxes without scanning the whole vault."""
    roots = list(DEFAULT_SOURCE_ROOTS)
    vault_root = Path(os.environ.get("AI_OS_VAULT_ROOT") or "/Volumes/Devarsh SSD/Obsidian memory ")
    data_root = Path(os.environ.get("AI_OS_DATA_ROOT") or "/Volumes/Devarsh SSD/AI OS Data")
    roots.extend(
        [
            ("vault_agent_outputs", vault_root / "ai memory" / "00 AI OS" / "Agent Outputs"),
            ("vault_research_outputs", vault_root / "ai memory" / "01 Research" / "AI Outputs"),
            ("shared_research_inbox", data_root / "research-inbox"),
        ]
    )
    configured = filter(
        None,
        os.environ.get("AI_OS_RESEARCH_EXTRA_ROOTS", "").split(os.pathsep),
    )
    for index, value in enumerate(configured, start=1):
        roots.append((f"configured_root_{index}", Path(value).expanduser()))

    deduplicated: list[tuple[str, Path]] = []
    seen: set[str] = set()
    for label, root in roots:
        normalized = str(root.expanduser())
        if normalized in seen:
            continue
        seen.add(normalized)
        deduplicated.append((label, Path(normalized)))
    return deduplicated


STANDALONE_FILES = [
    Path("/Users/devarshthakkar/Downloads/SJS_Enterprises_Institutional_Research_Report.pdf"),
    Path("/Users/devarshthakkar/Downloads/SJS_Enterprises_Executive_Summary.pdf"),
    Path("/Users/devarshthakkar/Downloads/SJS_Enterprises_Interactive_Dashboard.html"),
    Path("/Users/devarshthakkar/Downloads/SJS_Enterprises_Financial_Model.xlsx"),
    Path("/Users/devarshthakkar/Downloads/Zaggle Deep Dive.pdf"),
    Path("/Users/devarshthakkar/Downloads/Portfolio_Report.pdf"),
    Path("/Users/devarshthakkar/Downloads/Sanjana_Long Term_Report.pdf"),
    Path("/Users/devarshthakkar/Downloads/Sanjana_Long Term_Report_2025-09-17.pdf"),
    Path("/Users/devarshthakkar/Downloads/research_database.html"),
    Path("/Users/devarshthakkar/Downloads/research_dashboard.html"),
    Path("/Users/devarshthakkar/Downloads/research_db.html"),
]

SUPPORTED_SUFFIXES = {".pdf", ".html", ".htm", ".md", ".txt", ".csv", ".docx", ".xlsx", ".xlsm", ".json"}
SKIP_NAMES = {".DS_Store"}
SKIP_PREFIXES = (".~lock.",)


class TextHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.title: str | None = None
        self._in_title = False
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "title":
            self._in_title = True

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "title":
            self._in_title = False

    def handle_data(self, data: str) -> None:
        text = " ".join(html.unescape(data).split())
        if not text:
            return
        if self._in_title:
            self.title = text
        self.parts.append(text)


def clean_text(value: str, limit: int = 1200) -> str:
    cleaned = "".join(ch if ch >= " " or ch in "\n\t" else " " for ch in (value or ""))
    text = re.sub(r"\s+", " ", cleaned).strip()
    return text[:limit]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def root_files() -> Iterable[tuple[str, Path]]:
    seen: set[Path] = set()
    for label, root in configured_source_roots():
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path in seen or not path.is_file():
                continue
            seen.add(path)
            yield label, path
    for path in STANDALONE_FILES:
        if path.exists() and path not in seen:
            seen.add(path)
            yield "standalone_downloads", path


def is_supported(path: Path) -> bool:
    if path.name in SKIP_NAMES or path.name.startswith(SKIP_PREFIXES):
        return False
    return path.suffix.lower() in SUPPORTED_SUFFIXES


def detect_family(path: Path) -> str:
    name = path.name.lower()
    parent = str(path.parent).lower()
    if "dashboard" in name or "/dashboard" in parent or name.endswith(".html"):
        return "dashboard"
    if "model" in name or path.suffix.lower() in {".xlsx", ".xlsm"}:
        return "financial_model"
    if "summary" in name:
        return "executive_summary"
    if "audit" in name:
        return "source_audit"
    if path.suffix.lower() == ".json":
        return "data_pack"
    if "report" in name or path.suffix.lower() == ".pdf":
        return "research_report"
    if path.suffix.lower() == ".md":
        return "research_note"
    return "research_artifact"


def artifact_type_for(family: str) -> str:
    if family == "dashboard":
        return "ai_dashboard_output"
    if family == "financial_model":
        return "ai_model_output"
    return "ai_research_output"


def title_from_name(path: Path) -> str:
    name = path.stem.replace("_", " ").replace("-", " ")
    name = re.sub(r"\s+", " ", name).strip()
    return name[:1].upper() + name[1:] if name else path.name


def company_or_topic(path: Path) -> str:
    stem = title_from_name(path)
    stop = {
        "Equity", "Research", "Report", "Interactive", "Dashboard", "Institutional",
        "Executive", "Summary", "Financial", "Model", "Investment", "Valuation",
        "Workbook", "Scenario", "Pack", "Data", "Source", "Audit",
    }
    words = [word for word in stem.split() if word not in stop]
    return " ".join(words[:5]) if words else stem


def summarize_markdown(path: Path) -> tuple[str | None, str]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    heading = None
    for line in text.splitlines():
        if line.startswith("#"):
            heading = line.lstrip("#").strip()
            break
    return heading, clean_text(text)


def summarize_html(path: Path) -> tuple[str | None, str]:
    parser = TextHTMLParser()
    parser.feed(path.read_text(encoding="utf-8", errors="ignore"))
    return parser.title, clean_text(" ".join(parser.parts))


def summarize_pdf(path: Path) -> tuple[str | None, str]:
    if PdfReader is None:
        return None, "PDF text extraction unavailable: pypdf is not installed."
    try:
        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages[:3]:
            pages.append(page.extract_text() or "")
        metadata_title = None
        if reader.metadata and reader.metadata.title:
            metadata_title = str(reader.metadata.title)
            if metadata_title.strip().lower() in {"(anonymous)", "untitled"}:
                metadata_title = None
        return metadata_title, clean_text(" ".join(pages))
    except Exception as exc:  # keep inventory moving on damaged PDFs
        return None, f"PDF text extraction failed: {exc}"


def summarize_workbook(path: Path) -> tuple[str | None, str]:
    if load_workbook is None:
        return None, "Workbook inspection unavailable: openpyxl is not installed."
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        details = []
        for sheet in workbook.worksheets[:8]:
            details.append(f"{sheet.title}: {sheet.max_row} rows x {sheet.max_column} columns")
        return None, "; ".join(details)
    except Exception as exc:
        return None, f"Workbook inspection failed: {exc}"


def summarize_json(path: Path) -> tuple[str | None, str]:
    try:
        value = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
    except Exception as exc:
        return None, f"JSON parse failed: {exc}"
    if isinstance(value, dict):
        keys = ", ".join(list(value.keys())[:20])
        return None, f"JSON object keys: {keys}"
    if isinstance(value, list):
        return None, f"JSON list with {len(value)} items"
    return None, f"JSON scalar: {type(value).__name__}"


def summarize_text(path: Path) -> tuple[str | None, str]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    heading = None
    for line in text.splitlines():
        cleaned = line.strip()
        if cleaned:
            heading = cleaned[:120]
            break
    return heading, clean_text(text)


def summarize_csv(path: Path) -> tuple[str | None, str]:
    try:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            rows = []
            for index, row in enumerate(reader):
                rows.append(row[:12])
                if index >= 8:
                    break
        if not rows:
            return None, "CSV file is empty."
        header = rows[0]
        return None, clean_text(f"CSV columns: {', '.join(header)}. Sample rows: {rows[1:4]}")
    except Exception as exc:
        return None, f"CSV inspection failed: {exc}"


def summarize_docx(path: Path) -> tuple[str | None, str]:
    try:
        with zipfile.ZipFile(path) as archive:
            xml_text = archive.read("word/document.xml").decode("utf-8", errors="ignore")
        text = re.sub(r"<[^>]+>", " ", xml_text)
        text = html.unescape(text)
        return None, clean_text(text)
    except Exception as exc:
        return None, f"DOCX inspection failed: {exc}"


def summarize(path: Path) -> tuple[str | None, str]:
    suffix = path.suffix.lower()
    if suffix == ".md":
        return summarize_markdown(path)
    if suffix == ".txt":
        return summarize_text(path)
    if suffix == ".csv":
        return summarize_csv(path)
    if suffix == ".docx":
        return summarize_docx(path)
    if suffix in {".html", ".htm"}:
        return summarize_html(path)
    if suffix == ".pdf":
        return summarize_pdf(path)
    if suffix in {".xlsx", ".xlsm"}:
        return summarize_workbook(path)
    if suffix == ".json":
        return summarize_json(path)
    return None, ""


def artifact_record(root_label: str, path: Path) -> dict:
    stat = path.stat()
    family = detect_family(path)
    extracted_title, summary = summarize(path)
    mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return {
        "artifact_type": artifact_type_for(family),
        "title": extracted_title or title_from_name(path),
        "source_url": None,
        "local_path": str(path),
        "content_hash": sha256_file(path),
        "mime_type": mime_type,
        "sensitivity": "private",
        "metadata": {
            "root_label": root_label,
            "artifact_family": family,
            "company_or_topic": company_or_topic(path),
            "size_bytes": stat.st_size,
            "last_modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
            "summary": summary,
            "file_name": path.name,
            "suffix": path.suffix.lower(),
        },
    }


def run_psql(sql: str) -> str:
    command = [
        "docker", "exec", "-i", "ai_os_postgres", "psql",
        "-q", "-t", "-A", "-v", "ON_ERROR_STOP=1", "-U", "ai_os", "-d", "ai_os",
    ]
    completed = subprocess.run(command, input=sql, text=True, capture_output=True, check=False)
    if completed.returncode != 0:
        raise RuntimeError((completed.stderr or completed.stdout).strip())
    output = completed.stdout.strip()
    if not output:
        raise RuntimeError((completed.stderr or "psql completed without returning output").strip())
    return output


def sql_string(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def upsert_records(records: list[dict]) -> dict:
    payload = json.dumps(records)
    sql = f"""
WITH source AS (
    INSERT INTO core.source_systems (name, source_type, location, sensitivity, status, notes)
    VALUES (
        {sql_string(SOURCE_SYSTEM_NAME)},
        'local_ai_output_folder',
        {sql_string(SOURCE_SYSTEM_LOCATION)},
        'private',
        'indexed',
        'AI-generated reports, dashboards, models, and source packs discovered from bounded local folders.'
    )
    ON CONFLICT (name) DO UPDATE SET
        source_type = EXCLUDED.source_type,
        location = EXCLUDED.location,
        sensitivity = EXCLUDED.sensitivity,
        status = EXCLUDED.status,
        notes = EXCLUDED.notes
    RETURNING id
),
payload AS (
    SELECT * FROM jsonb_to_recordset($aios_json${payload}$aios_json$::jsonb) AS x(
        artifact_type TEXT,
        title TEXT,
        source_url TEXT,
        local_path TEXT,
        content_hash TEXT,
        mime_type TEXT,
        sensitivity TEXT,
        metadata JSONB
    )
),
upserted AS (
    INSERT INTO core.raw_artifacts (
        source_system_id, artifact_type, title, source_url, local_path,
        content_hash, mime_type, sensitivity, metadata
    )
    SELECT
        source.id,
        payload.artifact_type,
        payload.title,
        payload.source_url,
        payload.local_path,
        payload.content_hash,
        payload.mime_type,
        payload.sensitivity,
        payload.metadata
    FROM payload
    CROSS JOIN source
    ON CONFLICT (source_system_id, source_url, local_path, content_hash) DO UPDATE SET
        artifact_type = EXCLUDED.artifact_type,
        title = EXCLUDED.title,
        mime_type = EXCLUDED.mime_type,
        sensitivity = EXCLUDED.sensitivity,
        metadata = EXCLUDED.metadata,
        captured_at = now()
    RETURNING id
)
SELECT json_build_object(
    'records_seen', (SELECT count(*) FROM payload),
    'records_upserted', (SELECT count(*) FROM upserted)
)::text;
"""
    return json.loads(run_psql(sql))


def main() -> int:
    records = [artifact_record(label, path) for label, path in root_files() if is_supported(path)]
    records.sort(key=lambda row: (row["metadata"]["root_label"], row["title"], row["local_path"]))
    summary = upsert_records(records)
    summary["total_inventory_rows"] = int(run_psql("SELECT count(*) FROM research.v_ai_output_inventory;"))
    family_counts: dict[str, int] = {}
    for row in records:
        family = row["metadata"]["artifact_family"]
        family_counts[family] = family_counts.get(family, 0) + 1
    summary["family_counts"] = family_counts
    summary["source_roots"] = [str(root) for _, root in configured_source_roots()]
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
