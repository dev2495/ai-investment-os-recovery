#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import io
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from collections import Counter
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree

from runtime_storage import artifact_root


RUNTIME_ROOT = Path(os.environ.get("AI_OS_RUNTIME_ROOT") or Path(__file__).absolute().parents[1])
VAULT_ROOT = Path(os.environ.get("AI_OS_VAULT_ROOT") or RUNTIME_ROOT.parent)
INTAKE_ROOT = artifact_root("local_intake")
BUNDLED_PYTHON = Path(os.environ.get("AI_OS_LOCAL_ARTIFACT_PYTHON") or "/Volumes/Devarsh SSD/AI OS Data/venvs/research-ingest/bin/python3")
SOFFICE_BIN = Path("/Users/devarshthakkar/.cache/codex-runtimes/codex-primary-runtime/dependencies/bin/override/soffice")
PARSER_VERSION = "local_artifact_v1"
SUPPORTED_SUFFIXES = {".csv", ".tsv", ".xls", ".xlsx", ".pdf", ".docx", ".txt", ".md", ".json", ".html", ".htm", ".png", ".jpg", ".jpeg", ".webp"}
TABULAR_SUFFIXES = {".csv", ".tsv", ".xls", ".xlsx"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}
HEADER_KEYWORDS = {
    "amount", "brokerage", "buy", "client", "date", "exchange", "expiry",
    "instrument", "market", "net", "option", "price", "qty", "quantity",
    "rate", "script", "scrip", "sell", "settlement", "strike", "symbol",
    "time", "trade", "type",
}


def sql_literal(value: object) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\x00", "").replace("'", "''") + "'"


def sql_jsonb(value: object) -> str:
    return f"{sql_literal(json.dumps(value, sort_keys=True, default=str))}::jsonb"


def run_psql(sql: str) -> list[dict[str, Any]]:
    commands = [
        ["/opt/homebrew/opt/postgresql@16/bin/psql", "-h", "127.0.0.1", "-p", os.environ.get("AI_OS_POSTGRES_PORT", "54329"), "-q", "-t", "-A", "-v", "ON_ERROR_STOP=1", "-U", "ai_os", "-d", "ai_os"],
        ["/opt/homebrew/bin/docker", "exec", "-i", "ai_os_postgres", "psql", "-q", "-t", "-A", "-v", "ON_ERROR_STOP=1", "-U", "ai_os", "-d", "ai_os"],
    ]
    env = os.environ.copy()
    env.setdefault("PGPASSWORD", os.environ.get("AI_OS_POSTGRES_PASSWORD", "ai_os_local_dev_change_me"))
    errors: list[str] = []
    for command in commands:
        completed = subprocess.run(command, input=sql, text=True, capture_output=True, check=False, env=env)
        if completed.returncode == 0:
            return json.loads(completed.stdout.strip() or "[]")
        errors.append((completed.stderr or completed.stdout).strip())
    raise RuntimeError(" | ".join(errors))


def ensure_rich_runtime() -> None:
    try:
        import openpyxl  # noqa: F401
        import pypdf  # noqa: F401
        from PIL import Image  # noqa: F401
        return
    except Exception:
        pass
    if os.environ.get("AI_OS_LOCAL_ARTIFACT_RUNTIME_REEXEC") == "1" or not BUNDLED_PYTHON.is_file():
        raise RuntimeError("bundled document and spreadsheet runtime is unavailable")
    env = os.environ.copy()
    env["AI_OS_LOCAL_ARTIFACT_RUNTIME_REEXEC"] = "1"
    os.execve(str(BUNDLED_PYTHON), [str(BUNDLED_PYTHON), *sys.argv], env)


def allowed_roots() -> list[Path]:
    configured = [item.strip() for item in os.environ.get("AI_OS_LOCAL_ARTIFACT_ROOTS", "").split(os.pathsep) if item.strip()]
    defaults = [
        str(VAULT_ROOT),
        str(INTAKE_ROOT.parent),
        "/Volumes/Devarsh SSD",
        str(Path.home() / "Downloads"),
        str(Path.home() / "Desktop"),
        str(Path.home() / "Documents"),
        str(Path.home() / ".codex" / "attachments"),
    ]
    return [Path(value).expanduser().resolve() for value in configured + defaults]


def validate_source_path(raw_path: str, max_bytes: int) -> Path:
    path = Path(raw_path).expanduser().resolve(strict=True)
    if not path.is_file():
        raise ValueError(f"local artifact is not a file: {path}")
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError(f"unsupported local artifact format: {path.suffix or '<none>'}")
    if not any(path == root or root in path.parents for root in allowed_roots()):
        raise ValueError("local artifact must be inside the vault, external SSD, Downloads, Desktop, Documents, or Codex attachments")
    size = path.stat().st_size
    if size > max_bytes:
        raise ValueError(f"local artifact exceeds the {max_bytes // (1024 * 1024)} MB limit")
    return path


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_text(value: str) -> str:
    value = value.replace("\x00", " ")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def serializable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dt.date, dt.datetime, dt.time)):
        return value.isoformat()
    return str(value)


def inferred_type(values: Iterable[Any]) -> str:
    kinds = Counter()
    for value in values:
        if value in (None, ""):
            continue
        if isinstance(value, bool):
            kinds["boolean"] += 1
        elif isinstance(value, (int, float)):
            kinds["number"] += 1
        elif isinstance(value, (dt.date, dt.datetime)):
            kinds["date"] += 1
        else:
            text = str(value).strip()
            if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text.replace(",", "")):
                kinds["number"] += 1
            elif re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}(?:[ T].*)?", text):
                kinds["date"] += 1
            else:
                kinds["text"] += 1
    return kinds.most_common(1)[0][0] if kinds else "empty"


def header_score(row: list[Any]) -> int:
    score = 0
    for value in row:
        if value in (None, ""):
            continue
        tokens = set(re.findall(r"[a-z]+", str(value).lower()))
        score += len(tokens & HEADER_KEYWORDS)
    return score


def detect_header_index(rows: list[list[Any]]) -> int:
    candidates = [(header_score(row), index) for index, row in enumerate(rows[:50])]
    score, index = max(candidates, default=(0, 0))
    return index if score >= 3 else 0


def unique_headers(values: list[Any]) -> list[str]:
    seen: Counter[str] = Counter()
    headers: list[str] = []
    for index, value in enumerate(values):
        base = str(value).strip() if value not in (None, "") else f"column_{index + 1}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def used_width(row: list[Any]) -> int:
    for index in range(len(row) - 1, -1, -1):
        if row[index] not in (None, ""):
            return index + 1
    return 0


def occupied_column_regions(rows: list[list[Any]], gap_tolerance: int = 3) -> list[tuple[int, int]]:
    occupied = sorted({index for row in rows for index, value in enumerate(row) if value not in (None, "")})
    if not occupied:
        return [(0, 0)]
    regions: list[list[int]] = []
    for index in occupied:
        if not regions or index - regions[-1][-1] > gap_tolerance:
            regions.append([index])
        else:
            regions[-1].append(index)
    return [(region[0], region[-1] + 1) for region in regions]


def profile_rows(name: str, rows: list[list[Any]], truncated: bool = False) -> dict[str, Any]:
    width = max((used_width(row) for row in rows), default=0)
    normalized = [list(row[:width]) + [None] * (width - len(row[:width])) for row in rows]
    header_index = detect_header_index(normalized)
    raw_headers = normalized[header_index] if normalized else []
    headers = unique_headers(raw_headers)
    data_rows = normalized[header_index + 1:] if normalized else []
    columns = []
    for index, header in enumerate(headers):
        values = [row[index] for row in data_rows]
        columns.append({
            "name": header,
            "inferred_type": inferred_type(values[:500]),
            "non_null_count": sum(value not in (None, "") for value in values),
        })
    return {
        "name": name,
        "row_count": len(data_rows),
        "column_count": len(headers),
        "columns": columns,
        "sample_rows": [{headers[index]: serializable(row[index]) for index in range(len(headers))} for row in data_rows[:5]],
        "header_row_index": header_index,
        "preamble_rows": header_index,
        "truncated": truncated,
    }


def parse_delimited(path: Path) -> tuple[list[dict[str, Any]], str]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        sample = handle.read(65536)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
        rows: list[list[Any]] = []
        truncated = False
        for index, row in enumerate(csv.reader(handle, dialect)):
            if index >= 200001:
                truncated = True
                break
            rows.append(row)
    return [profile_rows(path.stem, rows, truncated)], "csv_stdlib"


def parse_xlsx(path: Path) -> tuple[list[dict[str, Any]], str]:
    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(path, read_only=True, data_only=True)
    profiles: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows: list[list[Any]] = []
        truncated = False
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if index >= 200001:
                truncated = True
                break
            rows.append(list(row))
        regions = occupied_column_regions(rows)
        for region_index, (start, end) in enumerate(regions, 1):
            sliced_rows = [row[start:end] for row in rows]
            region_name = worksheet.title if len(regions) == 1 else f"{worksheet.title} region {region_index}"
            profile = profile_rows(region_name, sliced_rows, truncated)
            profile.update({
                "source_sheet": worksheet.title,
                "source_column_start": start + 1,
                "source_column_end": end,
                "sparse_region": len(regions) > 1,
            })
            profiles.append(profile)
    workbook.close()
    return profiles, "openpyxl"


def parse_html_xls(path: Path) -> tuple[list[dict[str, Any]], str]:
    import pandas as pd  # type: ignore

    tables = pd.read_html(io.StringIO(path.read_text(encoding="utf-8", errors="replace")))
    profiles = []
    for index, frame in enumerate(tables):
        rows = [list(frame.columns)] + frame.where(frame.notna(), None).values.tolist()
        profiles.append(profile_rows(f"table_{index + 1}", rows))
    return profiles, "pandas_html"


def parse_xls(path: Path) -> tuple[list[dict[str, Any]], str]:
    signature = path.read_bytes()[:512].lstrip().lower()
    if signature.startswith(b"pk"):
        return parse_xlsx(path)
    if b"<html" in signature or b"<table" in signature:
        return parse_html_xls(path)
    if not SOFFICE_BIN.is_file():
        raise RuntimeError("legacy XLS conversion runtime is unavailable")
    temp_root = INTAKE_ROOT / "tmp"
    temp_root.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(dir=temp_root) as raw_temp:
        completed = subprocess.run(
            [str(SOFFICE_BIN), "--headless", "--convert-to", "xlsx", "--outdir", raw_temp, str(path)],
            capture_output=True,
            text=True,
            timeout=90,
            check=False,
        )
        converted = list(Path(raw_temp).glob("*.xlsx"))
        if completed.returncode != 0 or not converted:
            raise RuntimeError((completed.stderr or completed.stdout or "legacy XLS conversion failed").strip())
        profiles, _ = parse_xlsx(converted[0])
    return profiles, "libreoffice_openpyxl"


def parse_pdf(path: Path) -> tuple[str, int, str]:
    from pypdf import PdfReader  # type: ignore

    reader = PdfReader(str(path))
    text = clean_text("\n".join((page.extract_text() or "") for page in reader.pages))
    return text, len(reader.pages), "pypdf"


def parse_docx(path: Path) -> tuple[str, str]:
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("word/document.xml"))
    text = clean_text("\n".join(node.text or "" for node in root.iter() if node.tag.endswith("}t")))
    return text, "docx_xml"


def parse_text(path: Path) -> tuple[str, str]:
    text = path.read_text(encoding="utf-8", errors="replace")
    if path.suffix.lower() == ".json":
        parsed = json.loads(text)
        text = json.dumps(parsed, indent=2, ensure_ascii=False, default=str)
    return clean_text(text), "text_stdlib"


class _VisibleHTMLText(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.hidden_depth = 0
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style", "noscript", "template"}:
            self.hidden_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript", "template"} and self.hidden_depth:
            self.hidden_depth -= 1

    def handle_data(self, data: str) -> None:
        if not self.hidden_depth and data.strip():
            self.parts.append(data)


def parse_html(path: Path) -> tuple[str, str]:
    parser = _VisibleHTMLText()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    return clean_text("\n".join(parser.parts)), "html_parser_visible_text"


def parse_image(path: Path) -> tuple[int, int, str]:
    from PIL import Image  # type: ignore

    with Image.open(path) as image:
        width, height = image.size
        image_format = image.format or path.suffix.lstrip(".").upper()
    return width, height, f"pillow_{image_format.lower()}"


def infer_destination(path: Path, family: str) -> str:
    haystack = path.name.lower()
    if any(token in haystack for token in ["transaction", "tradebook", "contract_note"]):
        return "portfolio.transactions_staging"
    if any(token in haystack for token in ["holding", "portfolio"]):
        return "portfolio.holding_observations_staging"
    if "journal" in haystack or "option log" in haystack:
        return "trading.trade_journals_staging"
    if "equity" in haystack and "curve" in haystack:
        return "strategy.equity_curves_staging"
    if "strateg" in haystack or "backtest" in haystack:
        return "strategy.artifact_review_queue"
    if family == "image":
        return "ops.artifact_evidence"
    if family == "document":
        return "research.document_review_queue"
    return "core.integration_schema_mapping_required"


def persist(args: argparse.Namespace, path: Path, stored_path: Path, content_hash: str, profile: dict[str, Any]) -> dict[str, Any]:
    title = args.title.strip() or path.name
    source_reference = args.source_label.strip() or str(path)
    family = profile["artifact_family"]
    promotion_status = "needs_mapping" if family == "tabular" else "needs_review"
    destination = args.suggested_destination.strip() or infer_destination(path, family)
    ingestion_key = f"{PARSER_VERSION}:{content_hash}"
    run_key = args.run_key.strip() or f"local_artifact_{content_hash[:18]}"
    artifact_type = {"tabular": "operator_tabular_file", "document": "operator_document", "image": "operator_screenshot"}.get(family, "operator_local_file")
    task_title = f"Map or review local artifact: {title}"[:240]
    objective = f"Review checksum-backed {family} artifact, verify the suggested destination {destination}, and explicitly promote, remap, or exclude it. Never infer production investment rows without source validation."
    rows = run_psql(
        f"""
        WITH source_system AS (
            INSERT INTO core.source_systems (name, source_type, location, sensitivity, status, notes)
            VALUES ('operator local artifact intake', 'local_file_intake', {sql_literal(str(path.parent))}, {sql_literal(args.sensitivity)}, 'active',
                    'Operator-confirmed local files copied into immutable external AI OS artifact storage.')
            ON CONFLICT (name) DO UPDATE SET status='active', sensitivity=EXCLUDED.sensitivity
            RETURNING id
        ), artifact AS (
            INSERT INTO core.raw_artifacts (
                source_system_id, artifact_type, title, local_path,
                content_hash, mime_type, sensitivity, metadata
            )
            SELECT id, {sql_literal(artifact_type)}, {sql_literal(title)}, {sql_literal(str(stored_path))},
                   {sql_literal(content_hash)}, {sql_literal(profile.get('mime_type'))}, {sql_literal(args.sensitivity)},
                   {sql_jsonb({'original_path': source_reference, 'parser_version': PARSER_VERSION, 'operator_confirmed': True, 'file_size_bytes': path.stat().st_size})}
            FROM source_system
            ON CONFLICT (source_system_id, source_url, local_path, content_hash) DO UPDATE SET
                title=EXCLUDED.title, mime_type=EXCLUDED.mime_type,
                sensitivity=EXCLUDED.sensitivity,
                metadata=core.raw_artifacts.metadata || EXCLUDED.metadata,
                captured_at=now()
            RETURNING id
        ), upserted AS (
            INSERT INTO core.local_artifact_ingestions (
                ingestion_key, run_key, raw_artifact_id, source_path, stored_path,
                extracted_text_path, file_name, file_extension, artifact_family,
                mime_type, content_hash, file_size_bytes, parser_name, parser_version,
                status, promotion_status, suggested_destination, table_profiles,
                row_count, sheet_count, page_count, image_width, image_height,
                extracted_chars, text_preview, sensitivity, error_message,
                evidence, metadata, created_by
            )
            SELECT
                {sql_literal(ingestion_key)}, {sql_literal(run_key)}, id,
                {sql_literal(source_reference)}, {sql_literal(str(stored_path))},
                {sql_literal(profile.get('extracted_text_path'))}, {sql_literal(path.name)},
                {sql_literal(path.suffix.lower())}, {sql_literal(family)},
                {sql_literal(profile.get('mime_type'))}, {sql_literal(content_hash)}, {path.stat().st_size},
                {sql_literal(profile.get('parser_name'))}, {sql_literal(PARSER_VERSION)},
                {sql_literal(profile.get('status'))}, {sql_literal(promotion_status)}, {sql_literal(destination)},
                {sql_jsonb(profile.get('table_profiles') or [])},
                {profile.get('row_count') if profile.get('row_count') is not None else 'NULL'},
                {profile.get('sheet_count') if profile.get('sheet_count') is not None else 'NULL'},
                {profile.get('page_count') if profile.get('page_count') is not None else 'NULL'},
                {profile.get('image_width') if profile.get('image_width') is not None else 'NULL'},
                {profile.get('image_height') if profile.get('image_height') is not None else 'NULL'},
                {profile.get('extracted_chars') if profile.get('extracted_chars') is not None else 'NULL'},
                {sql_literal(profile.get('text_preview'))}, {sql_literal(args.sensitivity)},
                {sql_literal(profile.get('error_message'))},
                {sql_jsonb([{'source_path': source_reference, 'stored_path': str(stored_path), 'content_hash': content_hash}])},
                {sql_jsonb({'operator_confirmed': True, 'seed_data_allowed': False, 'automatic_promotion_allowed': False, 'intake_source_label': source_reference})},
                {sql_literal(args.actor)}
            FROM artifact
            ON CONFLICT (ingestion_key) DO UPDATE SET
                run_key=EXCLUDED.run_key, source_path=EXCLUDED.source_path,
                stored_path=EXCLUDED.stored_path,
                extracted_text_path=coalesce(EXCLUDED.extracted_text_path, core.local_artifact_ingestions.extracted_text_path),
                status=EXCLUDED.status, table_profiles=EXCLUDED.table_profiles,
                row_count=EXCLUDED.row_count, sheet_count=EXCLUDED.sheet_count,
                page_count=EXCLUDED.page_count, image_width=EXCLUDED.image_width,
                image_height=EXCLUDED.image_height, extracted_chars=EXCLUDED.extracted_chars,
                text_preview=EXCLUDED.text_preview, error_message=EXCLUDED.error_message,
                suggested_destination=CASE WHEN core.local_artifact_ingestions.promotion_status IN ('promoted','excluded')
                    THEN core.local_artifact_ingestions.suggested_destination ELSE EXCLUDED.suggested_destination END,
                promotion_status=CASE WHEN core.local_artifact_ingestions.promotion_status IN ('promoted','excluded')
                    THEN core.local_artifact_ingestions.promotion_status ELSE EXCLUDED.promotion_status END,
                evidence=core.local_artifact_ingestions.evidence || EXCLUDED.evidence,
                metadata=core.local_artifact_ingestions.metadata || EXCLUDED.metadata,
                last_seen_at=now(), seen_count=core.local_artifact_ingestions.seen_count+1,
                updated_at=now()
            RETURNING *
        )
        SELECT coalesce(json_agg(row_to_json(result_rows)), '[]'::json)::text
        FROM (
            SELECT id, ingestion_key, content_hash FROM upserted
        ) result_rows
        """
    )
    if not rows:
        raise RuntimeError("local artifact ingestion did not return an identity")
    ingestion_id = int(rows[0]["id"])
    run_psql(
        f"""
        WITH task_insert AS (
            INSERT INTO agent.tasks (
                title, objective, owner_agent, status, priority,
                approval_required, source_kind, source_ref, output_format, evidence
            )
            VALUES (
                {sql_literal(task_title)}, {sql_literal(objective)}, 'Data Steward',
                'queued', 'normal', false, 'core.local_artifact_ingestions',
                {sql_literal(str(ingestion_id))}, 'mapping_or_classification_decision',
                jsonb_build_array(jsonb_build_object('table','core.local_artifact_ingestions','id',{ingestion_id},'content_hash',{sql_literal(content_hash)}))
            )
            ON CONFLICT (owner_agent, source_kind, source_ref)
              WHERE status IN ('queued','in_progress','blocked')
                AND source_kind = 'core.local_artifact_ingestions'
                AND source_ref IS NOT NULL
            DO UPDATE SET
                title=EXCLUDED.title,
                objective=EXCLUDED.objective,
                evidence=agent.tasks.evidence || EXCLUDED.evidence,
                updated_at=now()
            RETURNING id
        ), linked AS (
            UPDATE core.local_artifact_ingestions ingestion
            SET task_id=(SELECT id FROM task_insert), updated_at=now()
            WHERE ingestion.id={ingestion_id}
            RETURNING ingestion.id
        )
        SELECT coalesce(json_agg(row_to_json(linked)), '[]'::json)::text FROM linked
        """
    )
    result_rows = run_psql(
        f"""
        SELECT coalesce(json_agg(row_to_json(result_rows)), '[]'::json)::text
        FROM (
            SELECT * FROM core.v_local_artifact_ingestion_queue
            WHERE id={ingestion_id}
        ) result_rows
        """
    )
    return result_rows[0] if result_rows else {}


def main() -> int:
    parser = argparse.ArgumentParser(description="Ingest an operator-confirmed local artifact into immutable external storage.")
    parser.add_argument("--local-path", required=True)
    parser.add_argument("--title", default="")
    parser.add_argument("--sensitivity", default="private", choices=["public", "internal", "private", "client_private", "restricted"])
    parser.add_argument("--suggested-destination", default="")
    parser.add_argument("--run-key", default="")
    parser.add_argument("--source-label", default="")
    parser.add_argument("--actor", default="Data Steward")
    parser.add_argument("--operator-confirmed", action="store_true")
    parser.add_argument("--max-mb", type=int, default=100)
    args = parser.parse_args()
    if not args.operator_confirmed:
        raise ValueError("operator confirmation is required before reading a local artifact")
    path = validate_source_path(args.local_path, max(1, min(args.max_mb, 200)) * 1024 * 1024)
    ensure_rich_runtime()
    content_hash = sha256_file(path)
    stored_dir = INTAKE_ROOT / content_hash[:2]
    stored_dir.mkdir(parents=True, exist_ok=True)
    stored_path = stored_dir / f"{content_hash}{path.suffix.lower()}"
    if not stored_path.exists():
        shutil.copy2(path, stored_path)
    mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    suffix = path.suffix.lower()
    profile: dict[str, Any] = {
        "artifact_family": "tabular" if suffix in TABULAR_SUFFIXES else "image" if suffix in IMAGE_SUFFIXES else "document",
        "mime_type": mime_type,
        "status": "registered",
        "parser_name": "binary_registry",
    }
    try:
        if suffix in {".csv", ".tsv"}:
            table_profiles, parser_name = parse_delimited(stored_path)
            profile.update(table_profiles=table_profiles, parser_name=parser_name, status="profiled")
        elif suffix == ".xlsx":
            table_profiles, parser_name = parse_xlsx(stored_path)
            profile.update(table_profiles=table_profiles, parser_name=parser_name, status="profiled")
        elif suffix == ".xls":
            table_profiles, parser_name = parse_xls(stored_path)
            profile.update(table_profiles=table_profiles, parser_name=parser_name, status="profiled")
        elif suffix == ".pdf":
            text, page_count, parser_name = parse_pdf(stored_path)
            profile.update(parser_name=parser_name, status="extracted", page_count=page_count)
            profile["extracted_text"] = text
        elif suffix == ".docx":
            text, parser_name = parse_docx(stored_path)
            profile.update(parser_name=parser_name, status="extracted")
            profile["extracted_text"] = text
        elif suffix in {".txt", ".md", ".json"}:
            text, parser_name = parse_text(stored_path)
            profile.update(parser_name=parser_name, status="extracted")
            profile["extracted_text"] = text
        elif suffix in {".html", ".htm"}:
            text, parser_name = parse_html(stored_path)
            profile.update(parser_name=parser_name, status="extracted")
            profile["extracted_text"] = text
        elif suffix in IMAGE_SUFFIXES:
            width, height, parser_name = parse_image(stored_path)
            profile.update(parser_name=parser_name, status="profiled", image_width=width, image_height=height)
    except Exception as exc:  # Persist parser failures as governed evidence.
        profile.update(status="failed", error_message=f"{type(exc).__name__}: {exc}")
    table_profiles = profile.get("table_profiles") or []
    profile["sheet_count"] = len({item.get("source_sheet") or item.get("name") for item in table_profiles}) if table_profiles else None
    logical_rows: dict[str, int] = {}
    for item in table_profiles:
        source_table = str(item.get("source_sheet") or item.get("name"))
        logical_rows[source_table] = max(logical_rows.get(source_table, 0), int(item.get("row_count") or 0))
    profile["row_count"] = sum(logical_rows.values()) if table_profiles else None
    extracted_text = str(profile.pop("extracted_text", ""))
    if extracted_text:
        text_path = stored_dir / f"{content_hash}.txt"
        text_path.write_text(extracted_text, encoding="utf-8")
        profile["extracted_text_path"] = str(text_path)
        profile["extracted_chars"] = len(extracted_text)
        profile["text_preview"] = extracted_text[:2000]
    result = persist(args, path, stored_path, content_hash, profile)
    print(json.dumps({"status": result.get("status"), "result": result}, indent=2, sort_keys=True, default=str))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001
        sys.stderr.write(f"{type(exc).__name__}: {exc}\n")
        raise SystemExit(1)
